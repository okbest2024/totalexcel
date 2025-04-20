import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font
import os
import datetime
import logging

# --- Configuration ---
SOURCE_FILE = '日报.xlsx'
TEMPLATE_FILE = '202504-例会模板.xlsx'
TEMPLATE_CATEGORY_COL = 3  # Column C in the template holds the categories
TARGET_COLUMN_HEADER = "本年累计" # Header text to find the target column
HEADER_ROW = 2 # Assuming the header is in row 2

# --- Logging Setup ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Helper Functions ---

def find_merged_cell_by_text(ws, target_text):
    """Finds the first merged cell range whose top-left cell contains the target text."""
    logging.info(f"查找包含文本 '{target_text}' 的合并单元格...")
    for merged_range in ws.merged_cells.ranges:
        top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        if top_left_cell.value and isinstance(top_left_cell.value, str) and target_text in top_left_cell.value:
            logging.info(f"找到合并单元格 '{target_text}' 在范围: {merged_range.coord}")
            return merged_range
    logging.warning(f"未找到包含文本 '{target_text}' 的合并单元格。")
    return None

def find_text_in_column_range(ws, target_text, min_row, max_row, min_col, max_col):
    """Finds the first cell containing target_text within specific row and column bounds."""
    logging.info(f"在行 {min_row}-{max_row}, 列 {get_column_letter(min_col)}-{get_column_letter(max_col)} 查找文本 '{target_text}'...")
    # Iterate downwards first, then across columns in the specified range
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
             # Check bounds just in case
            if r > ws.max_row or c > ws.max_column:
                continue
            cell = ws.cell(row=r, column=c)
            if cell.value and isinstance(cell.value, str) and target_text in cell.value:
                logging.info(f"找到文本 '{target_text}' 在单元格 {get_column_letter(c)}{r}")
                return r, c
    logging.warning(f"在指定范围内未找到文本 '{target_text}'。")
    return None, None

def find_first_non_empty_below(ws, start_row, col, max_rows_to_check=10):
    """Finds the first non-empty cell value directly below a starting cell."""
    logging.info(f"在 {get_column_letter(col)}{start_row} 下方查找第一个非空单元格 (最多检查 {max_rows_to_check} 行)...")
    for r in range(start_row + 1, min(start_row + 1 + max_rows_to_check, ws.max_row + 1)):
        cell_value = ws.cell(row=r, column=col).value
        logging.debug(f"检查单元格 {get_column_letter(col)}{r}: 值 = '{cell_value}'")
        # Consider 0 as a valid non-empty value
        if cell_value is not None and cell_value != "":
            logging.info(f"找到下方非空值 '{cell_value}' 在单元格 {get_column_letter(col)}{r}")
            return r, col, cell_value
    logging.warning(f"在 {get_column_letter(col)}{start_row} 下方未找到非空单元格。")
    return None, None, None

def find_template_category_row(template_ws, category_text):
    """Finds the row number for a specific category in the template."""
    logging.info(f"在模板文件第 {TEMPLATE_CATEGORY_COL} 列查找类别 '{category_text}'...")
    for row in range(1, template_ws.max_row + 1):
        cell_value = template_ws.cell(row=row, column=TEMPLATE_CATEGORY_COL).value
        if cell_value and isinstance(cell_value, str) and category_text in cell_value:
            logging.info(f"找到类别 '{category_text}' 在模板文件的第 {row} 行。")
            return row
    logging.warning(f"未在模板文件中找到类别 '{category_text}'。")
    return None

def find_column_index_by_header(ws, header_text, header_row):
    """Finds the column index (1-based) for a specific header text in the given header row."""
    logging.info(f"在模板文件第 {header_row} 行查找列标题 '{header_text}'...")
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value and isinstance(cell_value, str) and header_text.strip() == cell_value.strip():
            logging.info(f"找到列标题 '{header_text}' 在第 {col} 列 ({get_column_letter(col)})。")
            return col
    logging.error(f"错误：未能在模板文件的第 {header_row} 行找到列标题 '{header_text}'。")
    return None

def update_template(template_ws, target_row, target_col, value):
    """Updates the value in the target column of the specified row in the template, applying formatting."""
    if not target_row:
        logging.error("无法更新模板：未提供有效的类别行号。")
        return False
    if not target_col:
         logging.error("无法更新模板：未提供有效的目标列号。")
         return False

    try:
        target_cell = template_ws.cell(row=target_row, column=target_col)
        
        # --- Apply Formatting ---
        # 1. Round numeric values to 2 decimal places
        processed_value = value
        if isinstance(value, (int, float)):
            processed_value = round(value, 2)
        
        target_cell.value = processed_value
        logging.info(f"已更新模板文件单元格 {target_cell.coordinate} 的值为 '{processed_value}'")

        # 2. Set Alignment to Center
        target_cell.alignment = Alignment(horizontal='center', vertical='center') # Also center vertically
        
        # 3. Set Font Color to Black
        target_cell.font = Font(color='000000') # Black color code
        
        logging.info(f"已应用格式到单元格 {target_cell.coordinate}：居中对齐，黑色字体，数值保留两位小数")
        return True
    except Exception as e:
        logging.error(f"更新模板单元格 {get_column_letter(target_col)}{target_row} 或应用格式失败: {e}")
        return False


# --- Main Processing Logic ---

def main():
    logging.info("开始执行 Excel 数据提取与转移...")
    results_log = [] # To store findings for final output

    # --- File Existence Check ---
    if not os.path.exists(SOURCE_FILE):
        logging.error(f"错误：源文件 '{SOURCE_FILE}' 不存在。")
        return
    if not os.path.exists(TEMPLATE_FILE):
        logging.error(f"错误：模板文件 '{TEMPLATE_FILE}' 不存在。")
        return

    # --- Load Workbooks ---
    try:
        logging.info(f"加载源文件: {SOURCE_FILE}")
        source_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True) # data_only=True to get values, not formulas
        source_ws = source_wb.active
        logging.info(f"加载模板文件: {TEMPLATE_FILE}")
        template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
        template_ws = template_wb.active
        logging.info(f"源文件工作表: '{source_ws.title}', 模板文件工作表: '{template_ws.title}'")
    except Exception as e:
        logging.error(f"打开 Excel 文件时出错: {e}")
        return

    # --- Find Target Column Index --- Find this first!
    actual_target_col = find_column_index_by_header(template_ws, TARGET_COLUMN_HEADER, HEADER_ROW)
    if not actual_target_col:
        logging.error(f"无法继续处理，因为未能在模板第 {HEADER_ROW} 行找到目标列 '{TARGET_COLUMN_HEADER}'。")
        return # Stop processing if target column not found

    # --- Task 1-4: 代理买卖业务净收入 ---
    logging.info("\n--- 处理 代理买卖业务净收入 ---")
    xiaoji_range = find_merged_cell_by_text(source_ws, "小计（含股基、港股通、北交所、期权、债券、回购等）")
    if xiaoji_range:
        results_log.append(f"1. '小计...' 合并单元格位置: {xiaoji_range.coord}")
        # Search for '当年' within the columns of the '小计' merged cell, starting from its row
        dangnian_row, dangnian_col = find_text_in_column_range(
            source_ws, "当年",
            min_row=xiaoji_range.min_row,
            max_row=min(xiaoji_range.max_row + 5, source_ws.max_row), # Search a bit below as well
            min_col=xiaoji_range.min_col,
            max_col=xiaoji_range.max_col
        )
        if dangnian_row and dangnian_col:
            results_log.append(f"2. '当年' 单元格位置: {get_column_letter(dangnian_col)}{dangnian_row}")
            value_row, value_col, value = find_first_non_empty_below(source_ws, dangnian_row, dangnian_col)
            if value is not None:
                results_log.append(f"3. '当年' 下方数值: {value} (位置: {get_column_letter(value_col)}{value_row})")
                template_row = find_template_category_row(template_ws, "代理买卖业务净收入")
                if update_template(template_ws, template_row, actual_target_col, value):
                    results_log.append(f"4. 已将值 '{value}' 更新到模板 '代理买卖业务净收入' 对应行的 '{TARGET_COLUMN_HEADER}' 列 (第 {actual_target_col} 列)")
                else:
                    results_log.append(f"4. 更新模板 '代理买卖业务净收入' 失败 (行: {template_row}, 列: {actual_target_col})")
            else:
                 results_log.append(f"3. 未找到 '当年' 下方的非空值")
        else:
            results_log.append(f"2. 未在 '小计...' 列范围内找到 '当年' 单元格")
    else:
        results_log.append("1. 未找到 '小计...' 合并单元格")


    # --- Task 5: 最新双融余额规模（时点） ---
    logging.info("\n--- 处理 最新双融余额规模（时点） ---")
    dangri_range = find_merged_cell_by_text(source_ws, "当日余额（亿元）")
    if dangri_range:
        results_log.append(f"5a. '当日余额...' 合并单元格位置: {dangri_range.coord}")
        rongzi_row, rongzi_col = find_text_in_column_range(
            source_ws, "融资融券余额",
            min_row=dangri_range.min_row,
             max_row=min(dangri_range.max_row + 5, source_ws.max_row),
            min_col=dangri_range.min_col,
            max_col=dangri_range.max_col
        )
        if rongzi_row and rongzi_col:
            results_log.append(f"5b. '融资融券余额' (时点附近) 位置: {get_column_letter(rongzi_col)}{rongzi_row}")
            value_row, value_col, value = find_first_non_empty_below(source_ws, rongzi_row, rongzi_col)
            if value is not None:
                results_log.append(f"5c. '融资融券余额' 下方数值 (时点): {value} (位置: {get_column_letter(value_col)}{value_row})")
                template_row = find_template_category_row(template_ws, "最新双融余额规模（时点，万元）")
                if update_template(template_ws, template_row, actual_target_col, value):
                    results_log.append(f"5d. 已将值 '{value}' 更新到模板 '最新双融余额规模（时点...' 对应行的 '{TARGET_COLUMN_HEADER}' 列 (第 {actual_target_col} 列)")
                else:
                    results_log.append(f"5d. 更新模板 '最新双融余额规模（时点...' 失败 (行: {template_row}, 列: {actual_target_col})")
            else:
                results_log.append(f"5c. 未找到 '融资融券余额' 下方的非空值 (时点)")
        else:
             results_log.append(f"5b. 未在 '当日余额...' 列范围内找到 '融资融券余额'")
    else:
        results_log.append("5a. 未找到 '当日余额...' 合并单元格")

    # --- Task 6: 最新双融余额规模（日均） ---
    logging.info("\n--- 处理 最新双融余额规模（日均） ---")
    rijun_range = find_merged_cell_by_text(source_ws, "全年累计日均余额（亿元）")
    if rijun_range:
        results_log.append(f"6a. '全年累计日均余额...' 合并单元格位置: {rijun_range.coord}")
        # Assume '融资融券余额' is likely in the same row or nearby rows, within the column range
        rongzi_row, rongzi_col = find_text_in_column_range(
            source_ws, "融资融券余额",
            min_row=rijun_range.min_row,
            max_row=min(rijun_range.max_row + 5, source_ws.max_row),
            min_col=rijun_range.min_col,
            max_col=rijun_range.max_col
        )
        if rongzi_row and rongzi_col:
            results_log.append(f"6b. '融资融券余额' (日均附近) 位置: {get_column_letter(rongzi_col)}{rongzi_row}")
            value_row, value_col, value = find_first_non_empty_below(source_ws, rongzi_row, rongzi_col)
            if value is not None:
                results_log.append(f"6c. '融资融券余额' 下方数值 (日均): {value} (位置: {get_column_letter(value_col)}{value_row})")
                template_row = find_template_category_row(template_ws, "最新双融余额规模（日均，万元）")
                if update_template(template_ws, template_row, actual_target_col, value):
                    results_log.append(f"6d. 已将值 '{value}' 更新到模板 '最新双融余额规模（日均...' 对应行的 '{TARGET_COLUMN_HEADER}' 列 (第 {actual_target_col} 列)")
                else:
                    results_log.append(f"6d. 更新模板 '最新双融余额规模（日均...' 失败 (行: {template_row}, 列: {actual_target_col})")
            else:
                results_log.append(f"6c. 未找到 '融资融券余额' 下方的非空值 (日均)")
        else:
            results_log.append(f"6b. 未在 '全年累计日均余额...' 列范围内找到 '融资融券余额'")
    else:
        results_log.append("6a. 未找到 '全年累计日均余额...' 合并单元格")


    # --- Task 7: 新增有效户 ---
    logging.info("\n--- 处理 新增有效户 ---")
    youxiao_range = find_merged_cell_by_text(source_ws, "有效客户数（户）")
    if youxiao_range:
        results_log.append(f"7a. '有效客户数...' 合并单元格位置: {youxiao_range.coord}")
        # Search for '当年新增' within the columns of the '有效客户数' merged cell
        xinzeng_row, xinzeng_col = find_text_in_column_range(
            source_ws, "当年新增",
            min_row=youxiao_range.min_row,
            max_row=min(youxiao_range.max_row + 5, source_ws.max_row),
            min_col=youxiao_range.min_col,
            max_col=youxiao_range.max_col
        )
        # Exclude finding the longer string "当年新增客户期末流通净资产"
        if xinzeng_row and xinzeng_col and "期末流通净资产" not in source_ws.cell(xinzeng_row, xinzeng_col).value:
             results_log.append(f"7b. '当年新增' 单元格位置: {get_column_letter(xinzeng_col)}{xinzeng_row}")
             value_row, value_col, value = find_first_non_empty_below(source_ws, xinzeng_row, xinzeng_col)
             if value is not None:
                results_log.append(f"7c. '当年新增' 下方数值: {value} (位置: {get_column_letter(value_col)}{value_row})")
                template_row = find_template_category_row(template_ws, "新增有效户（户，不折算）") # Updated category text
                if update_template(template_ws, template_row, actual_target_col, value):
                     results_log.append(f"7d. 已将值 '{value}' 更新到模板 '新增有效户（户，不折算）' 对应行的 '{TARGET_COLUMN_HEADER}' 列 (第 {actual_target_col} 列)")
                else:
                    results_log.append(f"7d. 更新模板 '新增有效户（户，不折算）' 失败 (行: {template_row}, 列: {actual_target_col})")
             else:
                results_log.append(f"7c. 未找到 '当年新增' 下方的非空值")
        else:
            results_log.append(f"7b. 未在 '有效客户数...' 列范围内找到精确的 '当年新增' 单元格")
            # Maybe add a fallback search if the precise one fails? For now, just report not found.
    else:
        results_log.append("7a. 未找到 '有效客户数...' 合并单元格")


    # --- Task 8: 新增客户期末流通净资产 ---
    logging.info("\n--- 处理 新增客户期末流通净资产 ---")
    liutong_range = find_merged_cell_by_text(source_ws, "当年新增客户期末流通净资产（万元）")
    if liutong_range:
        results_log.append(f"8a. '当年新增客户期末流通净资产...' 合并单元格位置: {liutong_range.coord}")
        # Value is expected directly below the start of the merged cell range
        value_row, value_col, value = find_first_non_empty_below(source_ws, liutong_range.min_row, liutong_range.min_col)
        if value is not None:
            results_log.append(f"8b. '当年新增客户期末流通净资产...' 下方数值: {value} (位置: {get_column_letter(value_col)}{value_row})")
            template_row = find_template_category_row(template_ws, "新增客户期末流通净资产") # Using exact text from prompt
            if update_template(template_ws, template_row, actual_target_col, value):
                 results_log.append(f"8c. 已将值 '{value}' 更新到模板 '新增客户期末流通净资产' 对应行的 '{TARGET_COLUMN_HEADER}' 列 (第 {actual_target_col} 列)")
            else:
                 results_log.append(f"8c. 更新模板 '新增客户期末流通净资产' 失败 (行: {template_row}, 列: {actual_target_col})")
        else:
            results_log.append(f"8b. 未找到 '当年新增客户期末流通净资产...' 下方的非空值")
    else:
        results_log.append("8a. 未找到 '当年新增客户期末流通净资产...' 合并单元格")


    # --- Save Updated Template ---
    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{TEMPLATE_FILE.replace('.xlsx', '')}_已更新_{timestamp}.xlsx"
        logging.info(f"\n正在保存更新后的模板文件为: {output_filename}")
        template_wb.save(output_filename)
        logging.info("文件保存成功。")
        results_log.append(f"\n结果已保存到: {output_filename}")
    except Exception as e:
        logging.error(f"保存更新后的模板文件时出错: {e}")
        results_log.append(f"\n错误：无法保存更新后的文件。请检查文件是否被其他程序打开。")

    # --- Task 20: Output Results ---
    print("\n" + "="*20 + " 处理结果总结 " + "="*20)
    for line in results_log:
        print(line)
    print("="*55)
    logging.info("脚本执行完毕。")


if __name__ == "__main__":
    main() 